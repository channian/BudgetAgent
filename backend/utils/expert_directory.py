"""Resolve an expert's (Mandarin) name → email address.

Lookup order:
  1. A mapping file you maintain  (backend/data/expert_emails.xlsx or .csv)
  2. Active Directory displayName  (utils.ldap_lookup, optional fallback)

The mapping file is the authoritative source. It is cached in memory and
auto-reloaded whenever the file's modified-time changes, so you can update
the spreadsheet on the server without restarting the app.

Expected columns (header row, case-insensitive, Chinese or English):
    name / 姓名 / 專家 / 負責專家   → the expert's Mandarin name
    email / 信箱 / 電子郵件         → the email address
Only the first sheet of an .xlsx is read.
"""
import os
import csv
import logging

logger = logging.getLogger(__name__)

_DATA_DIR   = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
_XLSX_PATH  = os.path.join(_DATA_DIR, "expert_emails.xlsx")
_CSV_PATH   = os.path.join(_DATA_DIR, "expert_emails.csv")

_NAME_KEYS  = ("name", "姓名", "專家", "負責專家", "expert", "expert_name")
_EMAIL_KEYS = ("email", "信箱", "電子郵件", "mail", "e-mail")

# in-memory cache: {normalised_name: email}; reloaded when source mtime changes
_cache = {}
_cache_mtime = None
_cache_path = None


def _norm(s):
    """Normalise a name for matching: strip + drop internal whitespace."""
    return "".join((s or "").split()).lower()


def _pick_col(header, keys):
    """Return the index of the first header cell matching any key."""
    for i, cell in enumerate(header):
        c = (str(cell or "")).strip().lower()
        if c in keys:
            return i
    return None


def _load_csv(path):
    mapping = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return mapping
    header = rows[0]
    ni = _pick_col(header, _NAME_KEYS)
    ei = _pick_col(header, _EMAIL_KEYS)
    if ni is None or ei is None:
        logger.warning("expert_emails.csv: missing name/email column in header %r", header)
        return mapping
    for r in rows[1:]:
        if len(r) <= max(ni, ei):
            continue
        name, email = (r[ni] or "").strip(), (r[ei] or "").strip()
        if name and "@" in email:
            mapping[_norm(name)] = email
    return mapping


def _load_xlsx(path):
    mapping = {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed — cannot read %s", path)
        return mapping
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return mapping
    ni = _pick_col(header, _NAME_KEYS)
    ei = _pick_col(header, _EMAIL_KEYS)
    if ni is None or ei is None:
        logger.warning("expert_emails.xlsx: missing name/email column in header %r", header)
        return mapping
    for r in rows:
        if r is None or len(r) <= max(ni, ei):
            continue
        name = (str(r[ni]) if r[ni] is not None else "").strip()
        email = (str(r[ei]) if r[ei] is not None else "").strip()
        if name and "@" in email:
            mapping[_norm(name)] = email
    wb.close()
    return mapping


def _refresh_cache():
    """(Re)load the mapping file if it exists and has changed."""
    global _cache, _cache_mtime, _cache_path
    path = _XLSX_PATH if os.path.exists(_XLSX_PATH) else (
        _CSV_PATH if os.path.exists(_CSV_PATH) else None)
    if path is None:
        _cache, _cache_mtime, _cache_path = {}, None, None
        return
    mtime = os.path.getmtime(path)
    if path == _cache_path and mtime == _cache_mtime:
        return  # unchanged
    try:
        _cache = _load_xlsx(path) if path.endswith(".xlsx") else _load_csv(path)
        _cache_mtime, _cache_path = mtime, path
        logger.info("Loaded %d expert email mappings from %s", len(_cache), os.path.basename(path))
    except Exception as e:
        logger.warning("Failed to load expert email mapping %s: %s", path, e)


def resolve_email(expert_name: str):
    """Return the email for an expert name, or None. Mapping file first, then AD."""
    if not expert_name:
        return None

    # 1. mapping file
    _refresh_cache()
    email = _cache.get(_norm(expert_name))
    if email:
        return email

    # 2. AD fallback
    try:
        from utils.ldap_lookup import lookup_email_by_name
        return lookup_email_by_name(expert_name)
    except Exception as e:
        logger.debug("AD fallback failed for %r: %s", expert_name, e)
        return None
