import io, csv, json, hashlib, datetime
from flask import Blueprint, request, jsonify, send_file
from psycopg2.extras import Json as PgJson
from db import cursor as db_cursor, row_to_dict
from utils.audit import log as audit_log
from routes.auth import require_auth, current_user

budgets_bp = Blueprint("budgets", __name__)

PENDING_STATUSES   = ["AI_REVIEW", "EXPERT_REVIEW", "PENDING_ACTION"]
COMPLETED_STATUSES = ["CLOSED", "REJECTED"]

LOCK_TTL = 900  # seconds — 15-minute concurrency lock


def init_lock_columns():
    """Add locked_by / locked_at to budget_requests if not present."""
    try:
        with db_cursor(commit=True) as cur:
            cur.execute("""
                ALTER TABLE budget.budget_requests
                    ADD COLUMN IF NOT EXISTS locked_by  VARCHAR NULL,
                    ADD COLUMN IF NOT EXISTS locked_at  TIMESTAMP NULL
            """)
    except Exception:
        pass

# A case is "已審理" once the expert has acted on it (退件 / 補件 / 通過).
# Re-ingesting changed data for such a case must NOT overwrite history —
# it becomes a new 補送 case instead.
DECIDED_STATUSES = ("CLOSED", "REJECTED", "PENDING_ACTION")


def _case_signature(d):
    """Stable fingerprint of a case's meaningful content (ignores status /
    timestamps / expert decision). Used to detect 'same data re-scanned'."""
    amount = d.get("amount")
    try:
        amount = float(amount) if amount not in (None, "") else 0.0
    except (TypeError, ValueError):
        amount = 0.0
    ai = d.get("ai_result")
    if isinstance(ai, dict):
        ai_str = json.dumps(ai, sort_keys=True, ensure_ascii=False)
    else:
        ai_str = str(ai) if ai else ""
    parts = [
        (d.get("category")     or "").strip(),
        (d.get("sub_category") or "").strip(),
        (d.get("expert_name")  or "").strip(),
        (d.get("owner")        or "").strip(),
        f"{amount:.4f}",
        (d.get("ai_comment")   or "").strip(),
        ai_str,
    ]
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


def _next_supplement_name(cur, base):
    """Return the next free 補送 project name: X（補送）, X（補送2）, …"""
    candidate, n = f"{base}（補送）", 1
    while True:
        cur.execute(
            "SELECT 1 FROM budget.budget_requests WHERE project_name = %s",
            (candidate,),
        )
        if not cur.fetchone():
            return candidate
        n += 1
        candidate = f"{base}（補送{n}）"



def _notify_roles(roles, message):
    """Insert a notification row for every user whose role is in `roles`."""
    try:
        with db_cursor() as cur:
            cur.execute("SELECT id FROM budget.users WHERE role = ANY(%s)", (roles,))
            user_ids = [r["id"] for r in cur.fetchall()]
        if not user_ids:
            return
        with db_cursor(commit=True) as cur:
            for uid in user_ids:
                cur.execute(
                    "INSERT INTO budget.notifications (user_id, text) VALUES (%s, %s)",
                    (uid, message),
                )
    except Exception:
        pass


# ── List ──────────────────────────────────────────────────────────────
@budgets_bp.get("/budgets")
@require_auth
def list_budgets():
    scope    = request.args.get("scope", "pending")
    q        = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()

    statuses   = PENDING_STATUSES if scope == "pending" else COMPLETED_STATUSES
    conditions = ["status = ANY(%s)"]
    params     = [statuses]

    if q:
        conditions.append("(project_name ILIKE %s OR budget_no ILIKE %s)")
        like = f"%{q}%"
        params += [like, like]
    if category:
        conditions.append("category = %s")
        params.append(category)

    where = " AND ".join(conditions)
    try:
        with db_cursor() as cur:
            cur.execute(
                f"SELECT * FROM budget.budget_requests WHERE {where} ORDER BY id DESC",
                params,
            )
            rows = [row_to_dict(r) for r in cur.fetchall()]
    except Exception as e:
        return jsonify(error=str(e)), 500

    return jsonify(budgets=rows)


# ── Create / ingest ───────────────────────────────────────────────────
# Smart ingest used by both the manual "建立預算單" form and the RPA JSON drop.
# Behaviour when the same project_name already exists:
#   • data identical to any existing version → IGNORE   (kills re-scan spam)
#   • case still under review, data changed   → UPDATE   (pre-review fix)
#   • case already 已審理, data changed        → 補送 new case (keep history)
@budgets_bp.post("/budgets")
@require_auth
def create_budget():
    data = request.json or {}
    user = current_user()

    project = (data.get("project_name") or "").strip()
    if not project:
        return jsonify(error="缺少項目名稱（project_name）"), 400

    _, iso_week, _ = datetime.datetime.now().isocalendar()
    ai_obj = data.get("ai_result_obj")
    owner  = (data.get("owner") or user.get("name") or "").strip() or None
    status = "EXPERT_REVIEW" if ai_obj else "AI_REVIEW"

    incoming = {
        "category":     data.get("category"),
        "sub_category": data.get("sub_category"),
        "expert_name":  data.get("expert_name"),
        "owner":        owner,
        "amount":       data.get("amount", 0),
        "ai_comment":   data.get("ai_comment"),
        "ai_result":    ai_obj,
    }
    incoming_sig = _case_signature(incoming)

    try:
        # Gather the base case + all of its existing 補送 versions.
        with db_cursor() as cur:
            cur.execute(
                "SELECT * FROM budget.budget_requests "
                "WHERE project_name = %s OR project_name LIKE %s "
                "ORDER BY id DESC",
                (project, project.replace("%", r"\%").replace("_", r"\_") + "（補送%"),
            )
            existing = [row_to_dict(r) for r in cur.fetchall()]

        # 1) Identical data already present → ignore (re-scan no-op).
        for ex in existing:
            if _case_signature(ex) == incoming_sig:
                return jsonify(id=ex["id"], status=ex["status"], action="ignored",
                               project_name=ex["project_name"],
                               message="資料未變更，已存在相同案件，已略過。"), 200

        latest = existing[0] if existing else None

        # 2) Case still under review, data changed → update in place.
        if latest and latest["status"] not in DECIDED_STATUSES:
            with db_cursor(commit=True) as cur:
                cur.execute(
                    """UPDATE budget.budget_requests SET
                           category=%s, sub_category=%s, expert_name=%s,
                           owner=%s, amount=%s, ai_comment=%s, ai_result=%s,
                           status=%s, updated_at=NOW()
                       WHERE id=%s RETURNING *""",
                    (data.get("category"), data.get("sub_category"),
                     data.get("expert_name"), owner, data.get("amount", 0),
                     data.get("ai_comment"),
                     PgJson(ai_obj) if ai_obj else None, status, latest["id"]),
                )
                after = row_to_dict(cur.fetchone())
            audit_log(latest["id"], "UPDATE", user.get("name", "system"), latest, after)
            return jsonify(id=latest["id"], status=after["status"], action="updated",
                           project_name=after["project_name"],
                           message="案件尚未審理，已更新為最新資料。"), 200

        # 3) Decide the name: brand-new vs 補送 of an already-decided case.
        if latest and latest["status"] in DECIDED_STATUSES:
            with db_cursor() as cur:
                insert_name = _next_supplement_name(cur, project)
            note   = f"補送：原案件「{project}」(#{latest['id']}) 經審理後重新送件。"
            action = "supplemented"
        else:
            insert_name, note, action = project, None, "inserted"

        with db_cursor(commit=True) as cur:
            cur.execute(
                """INSERT INTO budget.budget_requests
                       (project_name, week, category, sub_category, expert_name,
                        owner, amount, ai_comment, ai_result, status, note)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   RETURNING id""",
                (insert_name, iso_week, data.get("category"),
                 data.get("sub_category"), data.get("expert_name"), owner,
                 data.get("amount", 0), data.get("ai_comment"),
                 PgJson(ai_obj) if ai_obj else None, status, note),
            )
            budget_id = cur.fetchone()["id"]
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "CREATE", user.get("name", "system"), None, data)
    if action == "supplemented":
        _notify_roles(["admin"],
            f"♻️ 收到補送案件「{insert_name}」(#{budget_id})，原案件「{project}」已審理過。")
    else:
        _notify_roles(["admin"],
            f"📝 {user.get('name','系統')} 建立了新案件「{insert_name}」(#{budget_id})。")
    return jsonify(id=budget_id, status=status, action=action,
                   project_name=insert_name), 201


# ── Get single ────────────────────────────────────────────────────────
@budgets_bp.get("/budgets/<int:budget_id>")
@require_auth
def get_budget(budget_id):
    try:
        with db_cursor() as cur:
            cur.execute(
                "SELECT * FROM budget.budget_requests WHERE id = %s",
                (budget_id,),
            )
            row = cur.fetchone()
    except Exception as e:
        return jsonify(error=str(e)), 500

    if not row:
        return jsonify(error="案件不存在"), 404
    return jsonify(budget=row_to_dict(row))


# ── Update (general fields) ───────────────────────────────────────────
@budgets_bp.put("/budgets/<int:budget_id>")
@require_auth
def update_budget(budget_id):
    data = request.json or {}
    user = current_user()

    allowed = {"budget_no", "expert_name", "owner", "amount", "category", "sub_category", "note",
               "expert_comment", "expert_decision"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify(error="無可更新欄位"), 400

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)

        set_clause = ", ".join(f"{k} = %s" for k in updates)
        with db_cursor(commit=True) as cur:
            cur.execute(
                f"UPDATE budget.budget_requests SET {set_clause} WHERE id = %s RETURNING *",
                [*updates.values(), budget_id],
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "UPDATE", user.get("name", "system"), before, after)
    _notify_roles(["admin"],
        f"✏️ {user.get('name','系統')} 更新了案件「{before.get('project_name','')}」(#{budget_id})，"
        f"修改欄位：{', '.join(updates.keys())}。")
    return jsonify(budget=after)


# ── Dispatch (AI_REVIEW → EXPERT_REVIEW) ─────────────────────────────
@budgets_bp.post("/budgets/<int:budget_id>/dispatch")
@require_auth
def dispatch_budget(budget_id):
    user = current_user()
    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可執行派發"), 403

    data        = request.json or {}
    budget_no   = (data.get("budget_no")   or "").strip() or None
    expert_name = (data.get("expert_name") or "").strip() or None

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)
        if before["status"] != "AI_REVIEW":
            return jsonify(error=f"只有 AI_REVIEW 狀態的案件可派發（目前：{before['status']}）"), 400

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET budget_no    = COALESCE(%s, budget_no),
                       expert_name  = COALESCE(%s, expert_name),
                       dispatch_date = NOW(),
                       status        = 'EXPERT_REVIEW',
                       updated_at    = NOW()
                   WHERE id = %s
                   RETURNING *""",
                (budget_no, expert_name, budget_id),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "DISPATCH", user.get("name", "system"), before, after)
    assigned = after.get("expert_name") or "未指定"
    _notify_roles(["admin"],
        f"📤 {user.get('name','系統')} 派發案件「{before['project_name']}」(#{budget_id})，"
        f"指派專家：{assigned}。")

    email_status = None
    if after.get("expert_name"):
        expert_name = after["expert_name"]

        # 1. In-app notification (existing logic)
        try:
            with db_cursor() as cur:
                cur.execute("SELECT id FROM budget.users WHERE name = %s", (expert_name,))
                eu = cur.fetchone()
            if eu:
                with db_cursor(commit=True) as cur:
                    cur.execute(
                        "INSERT INTO budget.notifications (user_id, text) VALUES (%s, %s)",
                        (eu["id"],
                         f"📋 案件「{before['project_name']}」(#{budget_id}) 已派發給您，請盡速審核。"),
                    )
        except Exception:
            pass

        # 2. Email lookup (mapping file → AD fallback) + actual email dispatch
        try:
            from utils.expert_directory import resolve_email
            from utils.email_service import send_dispatch_email
            expert_email = resolve_email(expert_name)
            if expert_email:
                ok = send_dispatch_email(
                    to_email=expert_email,
                    expert_name=expert_name,
                    project_name=before["project_name"],
                    budget_id=budget_id,
                    budget_no=after.get("budget_no"),
                    amount=after.get("amount"),
                    dispatch_date=after.get("dispatch_date"),
                )
                email_status = "sent" if ok else "failed"
            else:
                email_status = "no_email"
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Email dispatch error: %s", e)
            email_status = "error"

    return jsonify(budget=after, email_status=email_status)


# ── Reassign a dispatched case to a new expert (admin only) ──────────
@budgets_bp.post("/budgets/<int:budget_id>/reassign")
@require_auth
def reassign_budget(budget_id):
    user = current_user()
    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可重派案件"), 403

    data        = request.json or {}
    expert_name = (data.get("expert_name") or "").strip()
    reason      = (data.get("reason")      or "").strip()
    if not expert_name:
        return jsonify(error="請選擇新的負責專家"), 400
    if not reason:
        return jsonify(error="請填寫重派原因"), 400

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET expert_name   = %s,
                       dispatch_date = NOW(),
                       status        = 'EXPERT_REVIEW',
                       updated_at    = NOW()
                   WHERE id = %s
                   RETURNING *""",
                (expert_name, budget_id),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "REASSIGN", user.get("name", "system"), before, after)
    _notify_roles(["admin"],
        f"🔄 {user.get('name','系統')} 重派案件「{before['project_name']}」(#{budget_id})，"
        f"原因：{reason}，新指派：{expert_name}。")

    # In-app notification to new expert
    try:
        with db_cursor() as cur:
            cur.execute("SELECT id FROM budget.users WHERE name = %s", (expert_name,))
            eu = cur.fetchone()
        if eu:
            with db_cursor(commit=True) as cur:
                cur.execute(
                    "INSERT INTO budget.notifications (user_id, text) VALUES (%s, %s)",
                    (eu["id"],
                     f"📋 案件「{before['project_name']}」(#{budget_id}) 已重派給您，請盡速審核。"),
                )
    except Exception:
        pass

    # Email to new expert
    email_status = None
    try:
        from utils.expert_directory import resolve_email
        from utils.email_service import send_dispatch_email
        expert_email = resolve_email(expert_name)
        if expert_email:
            ok = send_dispatch_email(
                to_email=expert_email, expert_name=expert_name,
                project_name=before["project_name"], budget_id=budget_id,
                budget_no=after.get("budget_no"), amount=after.get("amount"),
                dispatch_date=after.get("dispatch_date"),
            )
            email_status = "sent" if ok else "failed"
        else:
            email_status = "no_email"
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Email reassign error: %s", e)
        email_status = "error"

    return jsonify(budget=after, email_status=email_status)


# ── Delete a case (admin only) ───────────────────────────────────────
@budgets_bp.delete("/budgets/<int:budget_id>")
@require_auth
def delete_budget(budget_id):
    user = current_user()
    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可刪除案件"), 403

    reason = ((request.json or {}).get("reason") or "").strip()
    if not reason:
        return jsonify(error="請提供刪除原因"), 400

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)

        with db_cursor(commit=True) as cur:
            # remove dependent audit logs first to satisfy FK
            cur.execute("DELETE FROM budget.audit_logs WHERE request_id = %s", (budget_id,))
            cur.execute("DELETE FROM budget.budget_requests WHERE id = %s", (budget_id,))
    except Exception as e:
        return jsonify(error=str(e)), 500

    _notify_roles(["admin"],
        f"🗑 {user.get('name','系統')} 刪除案件「{before['project_name']}」(#{budget_id})，"
        f"原因：{reason}")
    return jsonify(ok=True, deleted=budget_id)


# ── Concurrency lock ──────────────────────────────────────────────────
@budgets_bp.post("/budgets/<int:budget_id>/lock")
@require_auth
def acquire_lock(budget_id):
    """Try to acquire the 15-min edit lock for the review form."""
    user = current_user()
    if user.get("role") == "viewer":
        return jsonify(error="無編輯權限"), 403

    caller = user.get("name", "")
    try:
        with db_cursor() as cur:
            cur.execute("""
                SELECT locked_by, status,
                       CASE WHEN locked_at IS NOT NULL
                            THEN EXTRACT(EPOCH FROM (NOW() - locked_at))::int
                            ELSE NULL END AS lock_age_sec
                FROM budget.budget_requests WHERE id = %s
            """, (budget_id,))
            row = cur.fetchone()
    except Exception as e:
        return jsonify(error=str(e)), 500

    if not row:
        return jsonify(error="案件不存在"), 404
    r = row_to_dict(row)

    if r["status"] in COMPLETED_STATUSES:
        return jsonify(ok=False, reason="案件已結案")

    age       = r.get("lock_age_sec")          # seconds since lock was set
    holder    = (r.get("locked_by") or "").strip()
    expired   = (age is None) or (age > LOCK_TTL)

    if holder and holder != caller and not expired:
        return jsonify(ok=False, locked_by=holder,
                       expires_in=int(LOCK_TTL - age))

    # acquire / renew
    try:
        with db_cursor(commit=True) as cur:
            cur.execute(
                "UPDATE budget.budget_requests SET locked_by=%s, locked_at=NOW() WHERE id=%s",
                (caller, budget_id),
            )
    except Exception as e:
        return jsonify(error=str(e)), 500

    return jsonify(ok=True)


@budgets_bp.delete("/budgets/<int:budget_id>/lock")
@require_auth
def release_lock(budget_id):
    """Release the edit lock (only if held by the caller)."""
    caller = current_user().get("name", "")
    try:
        with db_cursor(commit=True) as cur:
            cur.execute("""
                UPDATE budget.budget_requests
                   SET locked_by = NULL, locked_at = NULL
                 WHERE id = %s AND locked_by = %s
            """, (budget_id, caller))
    except Exception:
        pass
    return jsonify(ok=True)


# ── Expert review (save comment + recommendation, no finalise) ───────
@budgets_bp.post("/budgets/<int:budget_id>/review")
@require_auth
def review_budget(budget_id):
    """Expert writes their comment + recommendation. Status stays EXPERT_REVIEW;
    the case then surfaces in the 待簽核 block for boss/admin to sign off."""
    user = current_user()
    if user.get("role") == "viewer":
        return jsonify(error="檢視者無法填寫專家評論"), 403

    data     = request.json or {}
    comment  = (data.get("comment") or "").strip() or None
    decision = data.get("decision")  # "通過" | "退件" | None
    if decision not in ("通過", "退件", None):
        return jsonify(error="建議處置必須是 通過/退件"), 400

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)

    # ── Concurrency lock check ────────────────────────────────────────
    # Reject if another user holds the edit lock and it hasn't expired.
        caller  = (user.get("name") or "").strip()
        role    = user.get("role", "viewer")
        holder  = (before.get("locked_by") or "").strip()
        age_row = None
        if holder and holder != caller and role not in ("admin",):
            with db_cursor() as cur:
                cur.execute("""
                    SELECT EXTRACT(EPOCH FROM (NOW() - locked_at))::int AS age
                    FROM budget.budget_requests WHERE id = %s
                """, (budget_id,))
                age_row = cur.fetchone()
            age = row_to_dict(age_row).get("age") if age_row else None
            if age is not None and age <= LOCK_TTL:
                return jsonify(
                    error=f"案件正在被「{holder}」編輯中，請稍後再試。"
                ), 423

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET expert_comment = %s, expert_decision = %s,
                       locked_by = NULL, locked_at = NULL, updated_at = NOW()
                   WHERE id = %s RETURNING *""",
                (comment, decision, budget_id),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "EXPERT_COMMENT", user.get("name", "system"), before, after)
    _notify_roles(["admin"],
        f"📝 {user.get('name','專家')} 已完成案件「{before['project_name']}」(#{budget_id}) 的專家評論，待簽核。")
    return jsonify(budget=after)


# ── Approve (CLOSED) ──────────────────────────────────────────────────
@budgets_bp.post("/budgets/<int:budget_id>/approve")
@require_auth
def approve_budget(budget_id):
    data    = request.json or {}
    user    = current_user()
    comment = data.get("comment", "")

    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可執行簽核"), 403

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)
        if before["status"] != "EXPERT_REVIEW":
            return jsonify(error=f"目前狀態 {before['status']} 無法執行核可"), 400

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET expert_decision = '通過',
                       expert_comment  = COALESCE(NULLIF(%s, ''), expert_comment),
                       sign_date       = NOW(),
                       cycle_time      = CASE
                           WHEN dispatch_date IS NOT NULL
                           THEN EXTRACT(DAY FROM (NOW() - dispatch_date))::INT
                           ELSE NULL
                       END,
                       status          = 'CLOSED',
                       locked_by       = NULL,
                       locked_at       = NULL,
                       updated_at      = NOW()
                   WHERE id = %s
                   RETURNING *""",
                (comment, budget_id),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "APPROVE", user.get("name", "system"), before, after)
    _notify_roles(["admin", "viewer"],
        f"✅ 案件「{before['project_name']}」(#{budget_id}) 已核准通過，由 {user.get('name','系統')} 簽核。")
    return jsonify(budget=after)


# ── Reject ────────────────────────────────────────────────────────────
@budgets_bp.post("/budgets/<int:budget_id>/reject")
@require_auth
def reject_budget(budget_id):
    data       = request.json or {}
    user       = current_user()
    comment    = data.get("comment", "")
    final      = bool(data.get("final", False))
    new_status = "REJECTED" if final else "PENDING_ACTION"

    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可執行簽核"), 403

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET expert_decision = '退件',
                       expert_comment  = COALESCE(NULLIF(%s, ''), expert_comment),
                       sign_date       = CASE WHEN %s THEN NOW() ELSE sign_date END,
                       status          = %s,
                       locked_by       = NULL,
                       locked_at       = NULL,
                       updated_at      = NOW()
                   WHERE id = %s
                   RETURNING *""",
                (comment, final, new_status, budget_id),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    action = "REJECT_FINAL" if final else "RETURN_FOR_SUPPLEMENT"
    audit_log(budget_id, action, user.get("name", "system"), before, after)
    if final:
        _notify_roles(["admin", "viewer"],
            f"❌ 案件「{before['project_name']}」(#{budget_id}) 已退件，由 {user.get('name','系統')} 審核。")
    else:
        _notify_roles(["admin", "viewer"],
            f"⚠ 案件「{before['project_name']}」(#{budget_id}) 退回補件，等待申請人補充資料。")
    return jsonify(budget=after)


# ── Batch sign (atomic) ──────────────────────────────────────────────
@budgets_bp.post("/budgets/batch-sign")
@require_auth
def batch_sign():
    """Sign off multiple cases in ONE transaction (all-or-nothing).

    Each case is approved if the expert recommended 通過, or finally
    rejected if the expert recommended 退件. If any single case fails
    validation, the whole batch rolls back and nothing is signed."""
    user = current_user()
    if user.get("role") != "admin":
        return jsonify(error="僅系統管理員可執行簽核"), 403

    ids = (request.json or {}).get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify(error="未提供要簽核的案件"), 400
    ids = [int(i) for i in ids]

    results = []   # (budget_id, before, after, action) — for audit/notify after commit
    try:
        # Single transaction: any raise inside → full rollback
        with db_cursor(commit=True) as cur:
            cur.execute(
                "SELECT * FROM budget.budget_requests WHERE id = ANY(%s) FOR UPDATE",
                (ids,),
            )
            rows = {r["id"]: row_to_dict(r) for r in cur.fetchall()}

            missing = [i for i in ids if i not in rows]
            if missing:
                raise ValueError(f"案件不存在：{missing}")

            bad = [i for i in ids if rows[i]["status"] != "EXPERT_REVIEW"]
            if bad:
                raise ValueError(f"案件狀態非專家審核中，無法簽核：{bad}")

            for i in ids:
                before = rows[i]
                if before.get("expert_decision") == "退件":
                    cur.execute(
                        """UPDATE budget.budget_requests
                           SET expert_decision = '退件', sign_date = NOW(),
                               status = 'REJECTED',
                               locked_by = NULL, locked_at = NULL, updated_at = NOW()
                           WHERE id = %s RETURNING *""",
                        (i,),
                    )
                    action = "REJECT_FINAL"
                else:
                    cur.execute(
                        """UPDATE budget.budget_requests
                           SET expert_decision = '通過', sign_date = NOW(),
                               cycle_time = CASE
                                   WHEN dispatch_date IS NOT NULL
                                   THEN EXTRACT(DAY FROM (NOW() - dispatch_date))::INT
                                   ELSE NULL END,
                               status = 'CLOSED',
                               locked_by = NULL, locked_at = NULL, updated_at = NOW()
                           WHERE id = %s RETURNING *""",
                        (i,),
                    )
                    action = "APPROVE"
                results.append((i, before, row_to_dict(cur.fetchone()), action))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=str(e)), 500

    # Post-commit side effects (non-fatal)
    operator = user.get("name", "system")
    for budget_id, before, after, action in results:
        try:
            audit_log(budget_id, action, operator, before, after)
            if action == "APPROVE":
                _notify_roles(["admin", "viewer"],
                    f"✅ 案件「{before['project_name']}」(#{budget_id}) 已核准通過，由 {operator} 簽核。")
            else:
                _notify_roles(["admin", "viewer"],
                    f"❌ 案件「{before['project_name']}」(#{budget_id}) 已退件，由 {operator} 審核。")
        except Exception:
            pass

    return jsonify(ok=True, signed=len(results))


# ── Resubmit ──────────────────────────────────────────────────────────
@budgets_bp.post("/budgets/<int:budget_id>/resubmit")
@require_auth
def resubmit_budget(budget_id):
    user = current_user()

    try:
        with db_cursor() as cur:
            cur.execute("SELECT * FROM budget.budget_requests WHERE id = %s", (budget_id,))
            before_row = cur.fetchone()
        if not before_row:
            return jsonify(error="案件不存在"), 404
        before = row_to_dict(before_row)
        if before["status"] != "PENDING_ACTION":
            return jsonify(error="只有待補件案件可重新遞交"), 400

        with db_cursor(commit=True) as cur:
            cur.execute(
                """UPDATE budget.budget_requests
                   SET status          = 'EXPERT_REVIEW',
                       expert_decision = NULL,
                       expert_comment  = NULL,
                       updated_at      = NOW()
                   WHERE id = %s
                   RETURNING *""",
                (budget_id,),
            )
            after = row_to_dict(cur.fetchone())
    except Exception as e:
        return jsonify(error=str(e)), 500

    audit_log(budget_id, "RESUBMIT", user.get("name", "system"), before, after)
    _notify_roles(["admin", "viewer"],
        f"🔄 案件「{before['project_name']}」(#{budget_id}) 已重新遞交，進入專家審核。")
    return jsonify(budget=after)


# ── Export (CSV / XLSX) ───────────────────────────────────────────────
EXPORT_COLUMNS = [
    ("budget_no",       "預算單號"),
    ("week",            "週數"),
    ("project_name",    "項目名稱"),
    ("category",        "類別"),
    ("sub_category",    "判定類別"),
    ("owner",           "預算負責人"),
    ("expert_name",     "負責專家"),
    ("amount",          "金額"),
    ("ai_comment",      "AI 初審評論"),
    ("expert_comment",  "專家複審評論"),
    ("status",          "狀態"),
    ("dispatch_date",   "派送日期"),
    ("sign_date",       "簽核日期"),
    ("cycle_time",      "Cycle Time"),
    ("note",            "備註"),
]


def _fetch_for_export(scope):
    statuses = PENDING_STATUSES if scope == "pending" else COMPLETED_STATUSES
    with db_cursor() as cur:
        cur.execute(
            "SELECT * FROM budget.budget_requests WHERE status = ANY(%s) ORDER BY id DESC",
            (statuses,),
        )
        return [row_to_dict(r) for r in cur.fetchall()]


@budgets_bp.get("/budgets/export")
@require_auth
def export_budgets():
    scope = request.args.get("scope", "pending")
    fmt   = request.args.get("format", "csv").lower()
    try:
        rows = _fetch_for_export(scope)
    except Exception as e:
        return jsonify(error=str(e)), 500

    keys    = [k for k, _ in EXPORT_COLUMNS]
    headers = [h for _, h in EXPORT_COLUMNS]
    stamp   = datetime.datetime.now().strftime("%Y%m%d_%H%M")

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            return jsonify(error="伺服器未安裝 openpyxl，請執行 pip install openpyxl"), 500
        wb = Workbook()
        ws = wb.active
        ws.title = "預算案件"
        ws.append(headers)
        for r in rows:
            ws.append([r.get(k, "") for k in keys])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"budget_{scope}_{stamp}.xlsx",
        )

    # CSV (UTF-8 BOM — Excel 開啟中文不亂碼)
    buf = io.StringIO()
    buf.write("﻿")
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(k, "") for k in keys])
    data = buf.getvalue().encode("utf-8")
    return send_file(
        io.BytesIO(data),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"budget_{scope}_{stamp}.csv",
    )


# ── Import (CSV / XLSX) ───────────────────────────────────────────────

# Aliases for the standard "new budget" import (→ AI_REVIEW)
IMPORT_ALIASES = {
    "project_name": ["項目名稱", "Project Name", "project_name", "project", "案件名稱"],
    "category":     ["類別", "category", "判定類別"],
    "owner":        ["預算負責人", "Owner", "owner", "負責人"],
    "amount":       ["金額", "amount", "金額 (NT$)"],
    "budget_no":    ["預算單號", "BudgetNo.", "BudgetNo", "Budget No.", "Budget No", "budget_no"],
}

# Aliases for the "completed history" import (→ CLOSED / REJECTED)
COMPLETED_IMPORT_ALIASES = {
    "project_name":    ["Project Name", "項目名稱", "案件名稱", "project_name"],
    "week":            ["週數(w)", "週數", "week", "Week"],
    # NOTE: Excel "類別" = 系統 (sub_category), NOT 大類別 (category)
    "category":        ["category", "判定類別"],
    "sub_category":    ["類別", "系統", "判定系統", "sub_category", "Sub Category"],
    # Excel 的「Owner」欄存的是負責專家，對應 expert_name（非 owner）
    "expert_name":     ["負責專家", "Owner", "Expert", "expert_name"],
    "budget_no":       ["BudgetNo.", "BudgetNo", "Budget No.", "預算單號", "budget_no"],
    "owner":           ["預算負責人", "owner", "負責人"],
    "amount":          ["金額 NT", "金額(NT$)", "金額(NT)", "金額", "amount", "金額 (NT$)"],
    "expert_comment":  ["專家評論", "expert_comment"],
    "expert_decision": ["審核處置", "expert_decision"],
    "dispatch_date":   ["派送日期", "dispatch_date"],
    "sign_date":       ["簽核日期", "sign_date"],
    "cycle_time":      ["Cycle time", "Cycle Time", "cycle_time", "CycleTime"],
    "note":            ["備註", "note"],
}


def _resolve_header(header, aliases=None):
    if aliases is None:
        aliases = IMPORT_ALIASES
    norm = {str(h).strip(): i for i, h in enumerate(header)}
    idx = {}
    for field, alts in aliases.items():
        for a in alts:
            if a in norm:
                idx[field] = norm[a]
                break
    return idx


def _parse_amount(v):
    if v is None:
        return 0
    try:
        return float(str(v).replace(",", "").replace("NT$", "").strip())
    except ValueError:
        return 0


def _parse_week(v, default):
    """Accept 'W21', '21', 21 → int week number."""
    if v is None:
        return default
    import re as _re
    m = _re.search(r'\d+', str(v))
    return int(m.group()) if m else default


def _parse_date(v):
    """Accept Excel datetime, date object, or string → date string or None."""
    if v is None:
        return None
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s  # return as-is; DB will reject invalid dates gracefully


def _parse_cycle(v):
    """Accept '3', '3天', '3.5' → int days or None."""
    if v is None:
        return None
    import re as _re
    m = _re.search(r'[\d.]+', str(v))
    if not m:
        return None
    try:
        return int(float(m.group()))
    except ValueError:
        return None


def _find_header_row(rows, max_scan=6):
    """Scan the first `max_scan` rows and return the index of the row that
    contains the most recognised column aliases (from either alias table).
    Falls back to row 0 if nothing matches."""
    all_aliases = set()
    for alts in IMPORT_ALIASES.values():
        all_aliases.update(alts)
    for alts in COMPLETED_IMPORT_ALIASES.values():
        all_aliases.update(alts)

    best_idx, best_hits = 0, 0
    for i, row in enumerate(rows[:max_scan]):
        hits = sum(1 for cell in row if str(cell or "").strip() in all_aliases)
        if hits > best_hits:
            best_hits, best_idx = hits, i
    return best_idx


def _load_sheet(f, sheet_name=None):
    """Load rows from an uploaded xlsx or csv file.
    Returns (header_row, data_rows) where all values are plain strings/scalars.
    Auto-detects which row contains the column headers (not necessarily row 1)."""
    name = f.filename.lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            raise ValueError("工作表是空的")
        hi = _find_header_row(data)
        return data[hi], data[hi + 1:]
    elif name.endswith(".csv"):
        rows = list(csv.reader(f.read().decode("utf-8-sig").splitlines()))
        if not rows:
            raise ValueError("檔案是空的")
        hi = _find_header_row(rows)
        return rows[hi], rows[hi + 1:]
    else:
        raise ValueError("僅支援 .csv 或 .xlsx 檔案")


# ── GET sheet names (probe endpoint) ─────────────────────────────────
@budgets_bp.post("/budgets/import/sheets")
@require_auth
def import_get_sheets():
    """Return the list of sheet names in an uploaded xlsx (or a placeholder for csv)."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="未收到檔案"), 400
    name = f.filename.lower()
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True, data_only=True)
            return jsonify(sheets=wb.sheetnames)
        except Exception as e:
            return jsonify(error=f"檔案解析失敗：{e}"), 500
    elif name.endswith(".csv"):
        return jsonify(sheets=["(單一工作表)"])
    else:
        return jsonify(error="僅支援 .csv 或 .xlsx 檔案"), 400


@budgets_bp.post("/budgets/import")
@require_auth
def import_budgets():
    user = current_user()
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="未收到檔案"), 400

    mode       = request.args.get("mode", "pending")   # "pending" | "completed"
    sheet_name = request.args.get("sheet", None)

    try:
        header, raw_rows = _load_sheet(f, sheet_name)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except Exception as e:
        return jsonify(error=f"檔案解析失敗：{e}"), 500

    if mode == "completed":
        return _import_completed(header, raw_rows, user)
    else:
        return _import_pending(header, raw_rows, user)


def _import_pending(header, raw_rows, user):
    """Original import logic: upsert into AI_REVIEW status."""
    idx = _resolve_header(header, IMPORT_ALIASES)
    if "project_name" not in idx:
        return jsonify(error="找不到「項目名稱 / Project Name」欄位，請確認標題列"), 400

    _, iso_week, _ = datetime.datetime.now().isocalendar()
    inserted, skipped, errors = 0, 0, []

    def cell(row, field):
        i = idx.get(field)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return None if v is None or str(v).strip() == "" else str(v).strip()

    try:
        with db_cursor(commit=True) as cur:
            for n, row in enumerate(raw_rows, start=2):
                project = cell(row, "project_name")
                if not project:
                    skipped += 1
                    continue
                cur.execute("SAVEPOINT _row")
                try:
                    cur.execute(
                        """INSERT INTO budget.budget_requests
                               (project_name, week, category, owner, amount, budget_no, status)
                           VALUES (%s, %s, %s, %s, %s, %s, 'AI_REVIEW')
                           ON CONFLICT (project_name) DO UPDATE SET
                               category  = EXCLUDED.category,
                               owner     = EXCLUDED.owner,
                               amount    = EXCLUDED.amount,
                               budget_no = COALESCE(EXCLUDED.budget_no, budget_requests.budget_no)
                           WHERE budget_requests.status NOT IN
                                 ('CLOSED', 'REJECTED', 'PENDING_ACTION')
                           RETURNING id""",
                        (project, iso_week, cell(row, "category"),
                         cell(row, "owner"), _parse_amount(cell(row, "amount")),
                         cell(row, "budget_no")),
                    )
                    row_result = cur.fetchone()
                    cur.execute("RELEASE SAVEPOINT _row")
                    if row_result:
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as row_err:
                    cur.execute("ROLLBACK TO SAVEPOINT _row")
                    errors.append(f"第 {n} 列：{row_err}")
    except Exception as e:
        return jsonify(error=str(e)), 500

    _audit_and_notify(user, inserted, skipped)
    return jsonify(inserted=inserted, skipped=skipped, errors=errors[:20])


def _import_completed(header, raw_rows, user):
    """Import historical completed records directly into CLOSED / REJECTED status."""
    idx = _resolve_header(header, COMPLETED_IMPORT_ALIASES)
    if "project_name" not in idx:
        # Help diagnose: show what columns we did recognise
        recognised = ", ".join(f"{k}←「{header[v]}」" for k, v in idx.items()) or "（無）"
        return jsonify(
            error=f"找不到「Project Name / 項目名稱」欄位，請確認標題列。"
                  f"已識別欄位：{recognised}"
        ), 400

    _, iso_week, _ = datetime.datetime.now().isocalendar()
    inserted, skipped, errors = 0, 0, []
    total_data_rows = sum(1 for r in raw_rows if any(v is not None and str(v).strip() for v in r))

    def cell(row, field):
        i = idx.get(field)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return None if v is None or str(v).strip() == "" else str(v).strip()

    def raw_cell(row, field):
        i = idx.get(field)
        if i is None or i >= len(row):
            return None
        return row[i]

    try:
        with db_cursor(commit=True) as cur:
            for n, row in enumerate(raw_rows, start=2):
                project = cell(row, "project_name")
                if not project:
                    skipped += 1
                    continue
                cur.execute("SAVEPOINT _row")
                try:
                    decision  = cell(row, "expert_decision") or ""
                    status    = "REJECTED" if "退件" in decision else "CLOSED"
                    exp_dec   = "退件" if "退件" in decision else ("通過" if decision else None)
                    wk        = _parse_week(raw_cell(row, "week"), iso_week)
                    amount    = _parse_amount(cell(row, "amount"))
                    d_date    = _parse_date(raw_cell(row, "dispatch_date"))
                    s_date    = _parse_date(raw_cell(row, "sign_date"))
                    cycle     = _parse_cycle(cell(row, "cycle_time"))

                    cur.execute(
                        """INSERT INTO budget.budget_requests
                               (project_name, week, category, sub_category, expert_name,
                                budget_no, owner, amount,
                                expert_comment, expert_decision, dispatch_date, sign_date,
                                cycle_time, note, status)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                           ON CONFLICT (project_name) DO UPDATE SET
                               week            = EXCLUDED.week,
                               category        = EXCLUDED.category,
                               sub_category    = COALESCE(EXCLUDED.sub_category, budget_requests.sub_category),
                               expert_name     = COALESCE(EXCLUDED.expert_name, budget_requests.expert_name),
                               budget_no       = COALESCE(EXCLUDED.budget_no, budget_requests.budget_no),
                               owner           = COALESCE(EXCLUDED.owner, budget_requests.owner),
                               amount          = COALESCE(NULLIF(EXCLUDED.amount,0), budget_requests.amount),
                               expert_comment  = EXCLUDED.expert_comment,
                               expert_decision = EXCLUDED.expert_decision,
                               dispatch_date   = EXCLUDED.dispatch_date,
                               sign_date       = EXCLUDED.sign_date,
                               cycle_time      = EXCLUDED.cycle_time,
                               note            = EXCLUDED.note,
                               status          = EXCLUDED.status
                           RETURNING id""",
                        (project, wk, cell(row, "category"), cell(row, "sub_category"),
                         cell(row, "expert_name"), cell(row, "budget_no"),
                         cell(row, "owner"), amount, cell(row, "expert_comment"),
                         exp_dec, d_date, s_date, cycle, cell(row, "note"), status),
                    )
                    row_result = cur.fetchone()
                    cur.execute("RELEASE SAVEPOINT _row")
                    if row_result:
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as row_err:
                    cur.execute("ROLLBACK TO SAVEPOINT _row")
                    errors.append(f"第 {n} 列：{row_err}")
    except Exception as e:
        return jsonify(error=str(e)), 500

    _audit_and_notify(user, inserted, skipped)
    return jsonify(
        inserted=inserted,
        skipped=skipped,
        total_rows=total_data_rows,
        errors=errors[:50],   # return up to 50 error lines for diagnostics
    )


def _audit_and_notify(user, inserted, skipped):
    try:
        audit_log(None, "IMPORT", user.get("name", "system"), None,
                  {"inserted": inserted, "skipped": skipped})
    except Exception:
        pass
    _notify_roles(["admin"],
        f"📥 {user.get('name','系統')} 匯入了 {inserted} 筆案件（略過 {skipped} 筆）。")


# ── Timeline ──────────────────────────────────────────────────────────
@budgets_bp.get("/budgets/<int:budget_id>/timeline")
@require_auth
def get_timeline(budget_id):
    try:
        with db_cursor() as cur:
            cur.execute(
                """SELECT id, action, operator, timestamp, diff_before, diff_after
                   FROM budget.audit_logs
                   WHERE request_id = %s
                   ORDER BY timestamp ASC""",
                (budget_id,),
            )
            rows = [row_to_dict(r) for r in cur.fetchall()]
    except Exception as e:
        return jsonify(error=str(e)), 500
    return jsonify(timeline=rows)
